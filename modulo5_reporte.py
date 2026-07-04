"""
Módulo 5: Reporte Final

Expone dos funciones públicas:
  actualizar_historial(causas) → APPEND a hoja CAUSAS de causas_ojv.xlsx
  generar_reporte(causas)      → crea Reporte_YYYY-MM-DD.xlsx

Input:  lista de causas con los campos del contrato M1 (incluye datos CBR)
Output 1: causas_ojv.xlsx (hoja CAUSAS actualizada — historial deduplicación)
Output 2: Reporte_YYYY-MM-DD.xlsx con 3 pestañas y formato condicional
"""

import os
import datetime
import logging

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

from config import CAUSAS_XLSX, BASE_DIR, REPORTES_DIR, SHEET_CAUSAS, MODELO_EXTRACCION, DEEPSEEK_PRECIO_INPUT_USD_POR_1M, DEEPSEEK_PRECIO_OUTPUT_USD_POR_1M

logging.basicConfig(level=logging.INFO, format="%(asctime)s [M5] %(message)s")
log = logging.getLogger(__name__)


# Orden geográfico norte a sur de las Cortes de Apelaciones de regiones.
# Las cortes de RM (Santiago y San Miguel) van en su propia pestaña y no
# aparecen aquí.
_ORDEN_CORTES_NORTE_SUR = [
    "C.A. de Arica",
    "C.A. de Iquique",
    "C.A. de Antofagasta",
    "C.A. de Copiapó",
    "C.A. de La Serena",
    "C.A. de Valparaíso",
    "C.A. de Rancagua",
    "C.A. de Talca",
    "C.A. de Chillán",
    "C.A. de Concepción",
    "C.A. de Temuco",
    "C.A. de Valdivia",
    "C.A. de Puerto Montt",
    "C.A. de Coyhaique",
    "C.A. de Punta Arenas",
]
_IDX_CORTE = {corte: i for i, corte in enumerate(_ORDEN_CORTES_NORTE_SUR)}


def _extraer_ordinal_tribunal(nombre: str) -> int:
    """Extrae el número ordinal de un nombre de tribunal para ordenar."""
    import re
    m = re.search(r"(\d+)\s*[°ºª]", nombre)
    if m:
        return int(m.group(1))
    m = re.search(r"^(\d+)\s+", nombre)
    if m:
        return int(m.group(1))
    return 9999  # tribunales sin ordinal van al final


# Prioridad de cortes para ordenamiento del reporte
_ORDEN_CORTES_REPORTE = {
    "C.A. de Santiago":   0,
    "C.A. de San Miguel": 1,
}


def _ordenar_regiones(causas: list[dict]) -> list[dict]:
    """
    Ordena causas por:
      1. Corte: C.A. de Santiago primero, luego C.A. de San Miguel, luego el resto
      2. Tribunal: orden ascendente por ordinal numérico (1°, 2°, …)
    """
    n_cortes_norte = len(_ORDEN_CORTES_NORTE_SUR)

    def key(c):
        corte = c.get("corte", "")
        # Prioridad de corte: Santiago=0, San Miguel=1, regiones por geografía (offset +2)
        idx_corte = _ORDEN_CORTES_REPORTE.get(corte)
        if idx_corte is None:
            idx_corte = _IDX_CORTE.get(corte, n_cortes_norte) + 2
        tribunal = c.get("tribunal", "")
        ordinal = _extraer_ordinal_tribunal(tribunal)
        return (idx_corte, ordinal, tribunal)

    return sorted(causas, key=key)


# ─────────────────────────────────────────────────────────────────
# Estilos reutilizables
# ─────────────────────────────────────────────────────────────────

_FILL_HEADER  = PatternFill("solid", fgColor="1F4E79")   # azul oscuro
_FILL_ODD     = PatternFill("solid", fgColor="F2F7FF")   # azul muy claro
_FILL_EVEN    = PatternFill("solid", fgColor="FFFFFF")   # blanco

_FONT_HEADER  = Font(bold=True, color="FFFFFF", size=11)
_FONT_BODY    = Font(size=10)
_FONT_BOLD    = Font(bold=True, size=10)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
_ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=False)

_BORDER_THIN  = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    right= Side(style="thin", color="CCCCCC"),
)


# ─────────────────────────────────────────────────────────────────
# Estructura de columnas
# ─────────────────────────────────────────────────────────────────

# (header_label, campo_dict, ancho_col, formato_num)
_COLUMNAS = [
    ("ROL",             "rol",                 9,  None),
    ("Año",             "año",                 6,  None),
    ("Corte",           "corte",              20,  None),
    ("Tribunal",        "tribunal",           30,  None),
    ("Demandante",      "demandante",         26,  None),
    ("Demandado",       "demandado",          26,  None),
    ("Dirección",       "direccion",          36,  None),
    ("Comuna",          "comuna",             16,  None),
    ("Fs.",             "fojas",               8,  None),
    ("CBR Motivo",      "cbr_motivo",         22,  None),
    ("Fechas Public.",  "fechas_publicacion", 18,  None),
    ("Fecha Remate",    "fecha_remate",       14,  None),
]


# ─────────────────────────────────────────────────────────────────
# Escritura de hoja de datos
# ─────────────────────────────────────────────────────────────────

def _escribir_hoja_datos(wb: openpyxl.Workbook, causas: list[dict], nombre: str, excluidos: list | None = None, descartados: list | None = None) -> None:
    """
    Crea una hoja con la lista de causas ordenadas, headers y formato condicional.
    """
    ws = wb.create_sheet(nombre)
    excluidos = excluidos or []
    descartados = descartados or []

    # ── Fila de encabezados ──
    for col_idx, (header, _, ancho, _fmt) in enumerate(_COLUMNAS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = _FILL_HEADER
        cell.font      = _FONT_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border    = _BORDER_THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"  # congelar fila de encabezados

    # ── Filas de datos ──
    for row_idx, causa in enumerate(causas, 2):
        fill = _FILL_ODD if row_idx % 2 == 1 else _FILL_EVEN

        for col_idx, (_, campo, _, fmt_num) in enumerate(_COLUMNAS, 1):
            valor = causa.get(campo, "")

            if valor is None:
                valor = ""
            # Corte: eliminar prefijo "C.A. de " para ahorrar espacio
            if campo == "corte" and isinstance(valor, str) and valor.startswith("C.A. de "):
                valor = valor[8:]

            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill      = fill
            cell.alignment = _ALIGN_RIGHT if fmt_num else _ALIGN_LEFT
            cell.font      = _FONT_BODY
            cell.border    = _BORDER_THIN

            if fmt_num and valor != "":
                cell.number_format = fmt_num

        ws.row_dimensions[row_idx].height = 16

    if not causas:
        return

    max_row = len(causas) + 1

    # ── Auto-filtro ──
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNAS))}{max_row}"

    # ── Tabla de transparencia CBR ──
    if max_row > 1:
        inicio_excl = max_row + 3  # 2 filas en blanco después de la tabla principal
    else:
        inicio_excl = 4

    # Título
    cell_tit = ws.cell(row=inicio_excl, column=1, value="PROPIEDADES NUEVAS EXCLUIDAS (dominio >= 2020)")
    cell_tit.font = Font(bold=True, color="FFFFFF", size=11)
    cell_tit.fill = PatternFill("solid", fgColor="8B0000")  # rojo oscuro
    ws.merge_cells(start_row=inicio_excl, start_column=1, end_row=inicio_excl, end_column=4)
    ws.row_dimensions[inicio_excl].height = 20

    # Headers
    headers_cbr = ["Tribunal", "ROL", "Año", "Año Dominio"]
    for ci, h in enumerate(headers_cbr, 1):
        cell = ws.cell(row=inicio_excl + 1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_THIN

    ws.row_dimensions[inicio_excl + 1].height = 18

    # Filas de datos
    if excluidos:
        for fi, exc in enumerate(excluidos):
            r_ex = inicio_excl + 2 + fi
            ws.cell(row=r_ex, column=1, value=exc.get("tribunal", "") or "")
            ws.cell(row=r_ex, column=2, value=exc.get("rol", "") or "")
            ws.cell(row=r_ex, column=3, value=exc.get("año", "") or "")
            ws.cell(row=r_ex, column=4, value=exc.get("cbr_anio", "") or "")
            for ci in range(1, 5):
                ws.cell(row=r_ex, column=ci).font = _FONT_BODY
                ws.cell(row=r_ex, column=ci).border = _BORDER_THIN
            ws.row_dimensions[r_ex].height = 16
    else:
        r_ex = inicio_excl + 2
        cell = ws.cell(row=r_ex, column=1, value="Sin exclusiones por CBR en esta edición")
        cell.font = _FONT_BODY
        ws.merge_cells(start_row=r_ex, start_column=1, end_row=r_ex, end_column=4)
        ws.row_dimensions[r_ex].height = 16

    # ── Tabla de causas descartadas por parámetros ──
    if excluidos:
        ultima_fila_cbr = inicio_excl + 2 + len(excluidos)
    else:
        ultima_fila_cbr = inicio_excl + 2
    inicio_desc = ultima_fila_cbr + 3  # 2 filas en blanco después de la tabla CBR

    cell_tit2 = ws.cell(row=inicio_desc, column=1, value="CAUSAS DESCARTADAS POR PARÁMETROS")
    cell_tit2.font = Font(bold=True, color="FFFFFF", size=11)
    cell_tit2.fill = PatternFill("solid", fgColor="8B0000")
    ws.merge_cells(start_row=inicio_desc, start_column=1, end_row=inicio_desc, end_column=5)
    ws.row_dimensions[inicio_desc].height = 20

    headers_desc = ["Tribunal", "ROL", "Año", "Motivo", "Monto (CLP)"]
    for ci, h in enumerate(headers_desc, 1):
        cell = ws.cell(row=inicio_desc + 1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_THIN
    ws.row_dimensions[inicio_desc + 1].height = 18

    if descartados:
        for fi, d in enumerate(descartados):
            r_d = inicio_desc + 2 + fi
            ws.cell(row=r_d, column=1, value=d.get("tribunal", "") or "")
            ws.cell(row=r_d, column=2, value=d.get("rol", "") or "")
            ws.cell(row=r_d, column=3, value=d.get("año", "") or "")
            ws.cell(row=r_d, column=4, value=d.get("motivo", "") or "")
            monto_val = d.get("monto")
            if monto_val:
                cell_m = ws.cell(row=r_d, column=5, value=monto_val)
                cell_m.number_format = '#,##0'
            else:
                cell_m = ws.cell(row=r_d, column=5, value="")
            for ci in range(1, 6):
                ws.cell(row=r_d, column=ci).font = _FONT_BODY
                ws.cell(row=r_d, column=ci).border = _BORDER_THIN
            ws.row_dimensions[r_d].height = 16
    else:
        r_d = inicio_desc + 2
        cell = ws.cell(row=r_d, column=1, value="Sin causas descartadas por parámetros en esta edición")
        cell.font = _FONT_BODY
        ws.merge_cells(start_row=r_d, start_column=1, end_row=r_d, end_column=5)
        ws.row_dimensions[r_d].height = 16


# ─────────────────────────────────────────────────────────────────
# Hoja Resumen
# ─────────────────────────────────────────────────────────────────

def _escribir_hoja_resumen(wb: openpyxl.Workbook, causas: list[dict], metricas: dict | None = None) -> None:
    """Crea la pestaña Resumen con estadísticas de la ejecución."""
    ws = wb.create_sheet("Resumen")
    ws.sheet_properties.tabColor = "1F4E79"
    metricas = metricas or {}

    # Anchos de columna
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 30

    fecha_hoy = datetime.date.today().strftime("%d/%m/%Y")

    # Contadores
    total = len(causas)

    # ── Helpers de escritura ──
    def titulo(row, texto):
        cell = ws.cell(row=row, column=1, value=texto)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_LEFT
        ws.merge_cells(f"A{row}:C{row}")
        ws.row_dimensions[row].height = 20

    def fila(row, label, valor, nota="", fill=None):
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=valor)
        c3 = ws.cell(row=row, column=3, value=nota)
        c1.font = c3.font = _FONT_BODY
        c2.font = _FONT_BOLD
        c2.alignment = _ALIGN_RIGHT
        if fill:
            c1.fill = c2.fill = c3.fill = fill
        ws.row_dimensions[row].height = 16

    def espacio(row):
        ws.row_dimensions[row].height = 8

    # ── Título ──
    r = 1
    ws.merge_cells(f"A{r}:C{r}")
    cell = ws.cell(row=r, column=1, value="RESUMEN — ANÁLISIS DE REMATES JUDICIALES")
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.fill = PatternFill("solid", fgColor="0D2D4F")
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[r].height = 28

    r += 1
    fila(r, "Fecha de procesamiento:", fecha_hoy)

    r += 1; espacio(r)
    r += 1; titulo(r, "TOTALES")
    r += 1; fila(r, "Total causas procesadas",   total)

    r += 1; espacio(r)
    r += 1; titulo(r, "MÉTRICAS DE EXTRACCIÓN")
    tokens_in = metricas.get("tokens_input", 0)
    tokens_out = metricas.get("tokens_output", 0)
    tokens_total = tokens_in + tokens_out
    r += 1; fila(r, "  Tokens entrada", f"{tokens_in:,}")
    r += 1; fila(r, "  Tokens salida",  f"{tokens_out:,}")
    r += 1; fila(r, "  Tokens total",   f"{tokens_total:,}")
    if MODELO_EXTRACCION == "deepseek":
        costo = (tokens_in / 1e6) * DEEPSEEK_PRECIO_INPUT_USD_POR_1M \
              + (tokens_out / 1e6) * DEEPSEEK_PRECIO_OUTPUT_USD_POR_1M
        r += 1; fila(r, "  Costo estimado", f"USD {costo:.4f}")
    else:
        r += 1; fila(r, "  Costo estimado", f"no calculado (motor={MODELO_EXTRACCION})")



# ─────────────────────────────────────────────────────────────────
# FUNCIÓN PÚBLICA 1: actualizar_historial
# ─────────────────────────────────────────────────────────────────

def actualizar_historial(causas: list[dict]) -> None:
    """
    Agrega las causas procesadas esta semana al historial acumulativo
    en la hoja CAUSAS de causas_ojv.xlsx.

    Solo hace APPEND — nunca borra filas existentes.
    Evita duplicar ROLes que ya estén en el historial.

    Args:
        causas: lista de dicts (output del pipeline completo)
    """
    log.info(f"Actualizando historial — {len(causas)} causa(s)")

    wb = openpyxl.load_workbook(CAUSAS_XLSX)

    if SHEET_CAUSAS not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_CAUSAS)
        ws.append(["ROL", "AÑO", "CORTE", "TRIBUNAL", "FECHA_PROCESADO"])
        log.warning(f"Hoja '{SHEET_CAUSAS}' no existía — se creó")
    else:
        ws = wb[SHEET_CAUSAS]
        # Agregar cabecera FECHA_PROCESADO si el sheet existente no la tiene
        if ws.cell(1, 5).value != "FECHA_PROCESADO":
            ws.cell(1, 5).value = "FECHA_PROCESADO"

    # Leer ROLes ya existentes para evitar duplicados
    roles_existentes: set[tuple] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            roles_existentes.add((str(row[0]).strip(), str(row[1]).strip()))

    fecha_hoy = datetime.date.today().isoformat()
    agregados = 0

    for c in causas:
        clave = (str(c.get("rol", "")).strip(), str(c.get("año", "")).strip())
        if not clave[0]:
            continue
        if clave in roles_existentes:
            continue

        ws.append([
            c.get("rol", ""),
            c.get("año", ""),
            c.get("corte", ""),
            c.get("tribunal", ""),
            fecha_hoy,
        ])
        roles_existentes.add(clave)
        agregados += 1

    wb.save(CAUSAS_XLSX)
    log.info(f"Historial actualizado: {agregados} causas nuevas agregadas")


# ─────────────────────────────────────────────────────────────────
# FUNCIÓN PÚBLICA 2: generar_reporte
# ─────────────────────────────────────────────────────────────────

def generar_reporte(causas: list[dict], metricas: dict | None = None) -> str:
    """
    Genera el reporte semanal Reporte_YYYY-MM-DD.xlsx con:
      - Pestaña "Regiones" (todas las causas, ordenadas geográficamente por ratio)
      - Pestaña "Resumen"  (estadísticas y notas)

    Args:
        causas: lista de dicts (output del pipeline completo)

    Returns:
        Ruta del archivo generado.
    """
    metricas = metricas or {}
    log.info(f"Generando reporte — {len(causas)} causa(s)")

    # Ordenar por corte (Santiago, San Miguel, regiones) y ordinal de tribunal
    regiones = _ordenar_regiones(causas)

    log.info(f"  Regiones: {len(regiones)}")

    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja vacía por defecto

    _escribir_hoja_datos(wb, regiones, "Regiones", excluidos=metricas.get("excluidos_cbr", []), descartados=metricas.get("descartados", []))
    _escribir_hoja_resumen(wb, causas, metricas=metricas)

    # Resumen queda al final (la hoja de detalle es la primera pestaña visible)
    # Resumen ya se creó después de Regiones, así que ya está al final.
    # Si hubiera más hojas, mover explícitamente:
    idx_resumen = wb.sheetnames.index("Resumen")
    if idx_resumen != len(wb.sheetnames) - 1:
        wb.move_sheet("Resumen", offset=len(wb.sheetnames) - 1 - idx_resumen)

    # Guardar — si el archivo del día ya existe y está bloqueado, agregar hora
    os.makedirs(REPORTES_DIR, exist_ok=True)
    fecha = datetime.date.today().isoformat()
    ruta  = os.path.join(REPORTES_DIR, f"Reporte_{fecha}.xlsx")
    if os.path.exists(ruta):
        try:
            import tempfile, shutil
            tmp = ruta + ".tmp"
            wb.save(tmp)
            shutil.move(tmp, ruta)
        except (PermissionError, OSError):
            hora = datetime.datetime.now().strftime("%H%M")
            ruta = os.path.join(REPORTES_DIR, f"Reporte_{fecha}_{hora}.xlsx")
            wb.save(ruta)
    else:
        wb.save(ruta)

    log.info(f"Reporte guardado: {ruta}")

    return ruta


# ─────────────────────────────────────────────────────────────────
# Standalone: genera reporte con datos sintéticos para probar formato
# ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    if "--demo" in sys.argv:
        # Genera un reporte de demostración sin correr el pipeline
        import random
        random.seed(42)

        comunas_r = ["Valparaíso", "Concepción", "Temuco", "La Serena", "Iquique"]

        causas_demo = []
        for i in range(1, 26):
            causas_demo.append({
                "rol":         str(30000 + i),
                "año":         str(random.randint(2015, 2023)),
                "corte":       f"C.A. de {random.choice(comunas_r)}",
                "tribunal":    f"{i}° Juzgado Civil de Concepción",
                "demandante":  random.choice(["Banco BCI", "Banco Itaú Chile", "Banco Santander"]),
                "demandado":   f"Deudor Demo {i}",
                "direccion":   f"Calle Demo {i * 100}",
                "comuna":      random.choice(comunas_r),
                "region_rm":   False,
            })

        print(f"Generando reporte demo con {len(causas_demo)} causas...")
        ruta = generar_reporte(causas_demo)
        print(f"Reporte creado: {ruta}")
