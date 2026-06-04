"""
modulo1_mercurio.py  —  Extractor El Mercurio Digital
======================================================
Scraper de avisos de remates judiciales de propiedades (sección 1616) desde
El Mercurio Digital, usando Playwright + Claude Text API.

CONTRATO DE DATOS  (interfaz con M2/M3/M5 — NO modificar)
----------------------------------------------------------
extraer_mercurio() retorna:
    list[ dict[str, Any] ] con claves exactas:
        rol, año, corte, tribunal, demandante, demandado,
        direccion, comuna, region_rm (siempre True)

Autor: generado automáticamente (Claude Sonnet 4.6)
Versión: 1.0  (2026-03-09)
"""

from __future__ import annotations

import asyncio
import json
import logging
import os
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from datetime import date, datetime


class EdicionNoDisponible(Exception):
    """La edición solicitada no está publicada en El Mercurio Digital."""
from pathlib import Path
from typing import Any

import anthropic
import openpyxl
from playwright.async_api import Page, async_playwright
from rapidfuzz import fuzz

from filtro_cbr import _CBR_ANIO_CORTE

# Tarifas DeepSeek para cálculo de costos
from config import (
    DEEPSEEK_PRECIO_INPUT_USD_POR_1M,
    DEEPSEEK_PRECIO_OUTPUT_USD_POR_1M,
)

# ---------------------------------------------------------------------------
# Importar config (credenciales / rutas / constantes)
# ---------------------------------------------------------------------------
try:
    from config import (
        ANTHROPIC_API_KEY,
        DEEPSEEK_API_KEY,
        MODELO_EXTRACCION,
        MERCURIO_USER,
        MERCURIO_PASS,
        MERCURIO_BASE_URL,
        CAUSAS_XLSX,          # ruta a causas_ojv.xlsx
    )
except ImportError as exc:
    raise SystemExit(
        "ERROR: config.py no encontrado o le faltan constantes requeridas.\n"
        "Asegúrate de definir: ANTHROPIC_API_KEY, MERCURIO_USER, MERCURIO_PASS, "
        "MERCURIO_BASE_URL, CAUSAS_XLSX\n"
        f"Detalle: {exc}"
    ) from exc

# ---------------------------------------------------------------------------
# Logger  (se configura con dual-logging en _setup_logging)
# ---------------------------------------------------------------------------
log = logging.getLogger("modulo1_mercurio")

_LOGS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")


class _LogFmt(logging.Formatter):
    """Formato: [HH:MM:SS] NIVEL — mensaje"""

    def format(self, record: logging.LogRecord) -> str:
        ts = time.strftime("%H:%M:%S", time.localtime(record.created))
        return f"[{ts}] {record.levelname} — {record.getMessage()}"


def _setup_logging() -> Path:
    """
    Configura dual-logging (consola + archivo) para modulo1_mercurio.
    Crea logs/ si no existe.
    Retorna la ruta del archivo de log creado.
    """
    os.makedirs(_LOGS_DIR, exist_ok=True)
    log_file = Path(_LOGS_DIR) / f"mercurio_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.log"

    # Limpiar handlers previos del logger de este módulo
    for h in log.handlers[:]:
        log.removeHandler(h)

    fmt = _LogFmt()

    console_h = logging.StreamHandler(
        open(sys.stdout.fileno(), mode="w", encoding="utf-8", closefd=False)
    )
    console_h.setFormatter(fmt)

    file_h = logging.FileHandler(str(log_file), encoding="utf-8")
    file_h.setFormatter(fmt)

    log.addHandler(console_h)
    log.addHandler(file_h)
    log.setLevel(logging.DEBUG)

    return log_file


# ---------------------------------------------------------------------------
# Estadísticas de ejecución
# ---------------------------------------------------------------------------
@dataclass
class _Stats:
    paginas_revisadas: int = 0
    paginas_con_1616: int = 0
    paginas_descartadas: int = 0
    pagina_parada: str = ""
    avisos_vision: int = 0
    avisos_post_filtro: int = 0
    causas_nuevas: int = 0
    seccion_utilizada: str = "F"
    tokens_input: int = 0
    tokens_output: int = 0
    excluidos_cbr: list = field(default_factory=list)  # [{tribunal, rol, año, cbr_anio}]
    descartados: list = field(default_factory=list)    # [{rol, año, tribunal, motivo, monto}]


def _log_resumen(stats: _Stats, *, dry_run: bool = False) -> None:
    """Imprime el bloque de resumen al final de la ejecución."""
    dr = " (dry run)" if dry_run else ""

    sec_label = stats.seccion_utilizada
    if sec_label == "B":
        sec_label = "B (fallback L-V)"
    elif sec_label == "D":
        sec_label = "D (clasificados fin de semana)"

    log.info("=" * 60)
    log.info("  RESUMEN EXTRACCIÓN MERCURIO DIGITAL%s", dr)
    log.info("=" * 60)
    log.info("  Sección utilizada       : %s", sec_label)
    log.info("  Páginas revisadas       : %d", stats.paginas_revisadas)
    log.info("  Descartadas (sin 1616)  : %d", stats.paginas_descartadas)
    log.info("  Conservadas (con 1616)  : %d", stats.paginas_con_1616)
    log.info("  Página de parada        : %s", stats.pagina_parada or "N/A")
    if dry_run:
        log.info("  Avisos Vision           : — (dry run)")
        log.info("  Post-filtro             : — (dry run)")
        log.info("  Nuevos (no dup)         : — (dry run)")
    else:
        log.info("  Avisos Vision           : %d", stats.avisos_vision)
        log.info("  Post-filtro             : %d", stats.avisos_post_filtro)
        log.info("  Nuevos (no dup)         : %d", stats.causas_nuevas)
    log.info("  Tokens entrada          : %d", stats.tokens_input)
    log.info("  Tokens salida           : %d", stats.tokens_output)
    if MODELO_EXTRACCION == "deepseek":
        costo = (stats.tokens_input / 1e6) * DEEPSEEK_PRECIO_INPUT_USD_POR_1M \
              + (stats.tokens_output / 1e6) * DEEPSEEK_PRECIO_OUTPUT_USD_POR_1M
        log.info("  Costo estimado          : USD %.4f (motor=deepseek)", costo)
    else:
        log.info("  Costo no calculado (motor=%s)", MODELO_EXTRACCION)
    log.info("=" * 60)


# ---------------------------------------------------------------------------
# Constantes internas
# ---------------------------------------------------------------------------
_UMBRAL_FUZZY_TRIBUNAL: int = 80          # RapidFuzz token_set_ratio threshold
_MAX_PAGINAS: int = 15                    # Tope de seguridad: máximas páginas a revisar
_CANVAS_HD_UMBRAL: int = 1800            # canvas.width > este valor → HD activo
_SECCIONES_MENORES = {"1611", "1612", "1613", "1614", "1615"}
_CORTES_RM = {"C.A. de Santiago", "C.A. de San Miguel"}
_BANCOS_ESTADO = {"banco estado", "banco del estado"}
_COMUNAS_EXCLUIDAS = {"estación central", "estacion central"}

MAX_WORKERS = 3  # nº de páginas procesadas en paralelo contra la API
# Puede subirse para medir el techo del tier. Si aparecen truncamientos
# o errores 429, bajarlo.

# Patrón para detectar recuadro de redirección a otra sección
# Ejemplo: "MÁS AVISOS ECONÓMICOS CLASIFICADOS EN PÁG. C 8"
_REDIRECT_PATTERN = re.compile(
    r"MÁS\s+AVISOS\s+ECON[OÓ]MICOS\s*CLASIFICADOS\s+EN\s+PÁG\.?\s*([A-Z])\s+(\d+)",
    re.IGNORECASE,
)

# ---------------------------------------------------------------------------
# Prompt para Claude Vision API
# ---------------------------------------------------------------------------
PROMPT_EXTRACCION = """Analiza este texto extraído de la sección "1616 — Remates de propiedades" del diario El Mercurio.

El texto viene del visor PDF y puede tener palabras cortadas por guiones de salto de línea (ej: "Juzga-\ndo" = "Juzgado", "San-\ntiago" = "Santiago"). Reconstrúyelas.

Extrae TODOS los avisos de remates de propiedades. Para cada aviso, devuelve:

- "rol": número del ROL de la causa (solo el número, sin "C-"). Formato: "XXXXX"
- "año": año del ROL (los últimos 4 dígitos después del último guión en el formato C-XXXXX-YYYY). Formato: "YYYY"
- "tribunal": nombre completo del tribunal (ej: "1° Juzgado Civil de Santiago")
- "demandante": nombre del demandante/ejecutante (banco o persona)
- "demandado": nombre del demandado/ejecutado
- "direccion": dirección completa del inmueble rematado
- "comuna": comuna donde se ubica el inmueble
- "fecha_remate": fecha del remate si aparece (formato DD/MM/YYYY)
- "año_inscripcion_dominio": año en que el dominio fue inscrito en el Conservador de Bienes Raíces (formato "YYYY"). Búscalo cerca de frases como "inscrito a fojas", "Registro de Propiedad", "Conservador", "año". Si no aparece, null.
- "fojas": el número de fojas de la inscripción del dominio (ej: "1234" o "1234 vta."). Suele aparecer como "fojas 1234", "a Fs. 1234", "inscrito a fojas 1234 N°...". Si no aparece, null.
REGLAS:
1. NO inventar datos. Si un campo no es identificable en el texto, devolver null.
2. El ROL siempre aparece como "Rol C-XXXXX-YYYY" o "Rol: C-XXXXX-YYYY" o "rol C-XXXXX-YYYY". El número es XXXXX y el año es YYYY.
3. El tribunal es el JUZGADO que ordena el remate, NO la dirección del tribunal.
4. SOLO extraer avisos de la sección 1616 (Remates de propiedades). Ignorar secciones 1611, 1612, 1615 u otras.
5. Si un aviso está cortado (al inicio o final del texto), extraer lo visible con campos faltantes como null.

Responde ÚNICAMENTE con un JSON array válido. Sin texto explicativo, sin markdown, sin comentarios. Solo JSON puro."""


# ===========================================================================
# FUNCIONES DE POST-PROCESAMIENTO (adaptadas del proyecto base)
# ===========================================================================

def _limpiar_tribunal(nombre: str | None) -> str | None:
    """
    Normaliza el nombre de un tribunal:
    - Une guiones silábicos (ej: "Juzga-\ndo" → "Juzgado")
    - Elimina fragmentos de dirección física
    - Normaliza mayúsculas/minúsculas
    """
    if not nombre:
        return None
    # Unir palabras partidas por guion al final de línea
    texto = re.sub(r"-\s*\n\s*", "", nombre)
    # Limpiar saltos de línea y espacios múltiples
    texto = re.sub(r"\s+", " ", texto).strip()
    # Eliminar texto entre paréntesis (a veces contiene dirección)
    texto = re.sub(r"\(.*?\)", "", texto).strip()
    # Capitalización básica: primera letra mayúscula en cada token relevante
    # (No alterar ordinales: 1°, 2°, etc.)
    texto = re.sub(r"\s{2,}", " ", texto)
    return texto


# ── Normalizador de ordinal de tribunal ──
def _normalizar_ordinal_tribunal(nombre: str | None) -> str | None:
    """
    Normaliza el ordinal escrito de un nombre de tribunal a formato "N°".
    Ej: "Primer Juzgado Civil de Santiago" → "1° Juzgado Civil de Santiago".
    Si ya es numérico o no tiene ordinal reconocible, deja intacto.
    """
    if not nombre:
        return nombre

    # Si ya tiene formato numérico con °, no tocar
    if re.search(r'\b\d+\s*[°º]', nombre):
        return nombre

    # Diccionario plano 1-30: clave → reemplazo "N°"
    # Incluye todas las variantes de género, apócope, con/sin tilde.
    ORDINALES_MAP = {
        # ── 30 ──
        "trigésimo": "30°", "trigesimo": "30°",
        "trigésima": "30°", "trigesima": "30°",
        # ── 29 ──
        "vigésimo noveno":  "29°", "vigesimo noveno":  "29°",
        "vigésima novena":  "29°", "vigesima novena":  "29°",
        # ── 28 ──
        "vigésimo octavo":  "28°", "vigesimo octavo":  "28°",
        "vigésima octava":  "28°", "vigesima octava":  "28°",
        # ── 27 ──
        "vigésimo séptimo": "27°", "vigesimo septimo": "27°",
        "vigésima séptima": "27°", "vigesima septima": "27°",
        # ── 26 ──
        "vigésimo sexto":   "26°", "vigesimo sexto":   "26°",
        "vigésima sexta":   "26°", "vigesima sexta":   "26°",
        # ── 25 ──
        "vigésimo quinto":  "25°", "vigesimo quinto":  "25°",
        "vigésima quinta":  "25°", "vigesima quinta":  "25°",
        # ── 24 ──
        "vigésimo cuarto":  "24°", "vigesimo cuarto":  "24°",
        "vigésima cuarta":  "24°", "vigesima cuarta":  "24°",
        # ── 23 ──
        "vigésimo tercero": "23°", "vigesimo tercero": "23°",
        "vigésimo tercer":  "23°", "vigesimo tercer":  "23°",
        "vigésima tercera": "23°", "vigesima tercera": "23°",
        # ── 22 ──
        "vigésimo segundo": "22°", "vigesimo segundo": "22°",
        "vigésima segunda": "22°", "vigesima segunda": "22°",
        # ── 21 ──
        "vigésimo primero": "21°", "vigesimo primero": "21°",
        "vigésimo primer":  "21°", "vigesimo primer":  "21°",
        "vigésima primera": "21°", "vigesima primera": "21°",
        # ── 20 ──
        "vigésimo": "20°", "vigesimo": "20°",
        "vigésima": "20°", "vigesima": "20°",
        # ── 19 ──
        "décimo noveno":    "19°", "decimo noveno":    "19°",
        "décima novena":    "19°", "decima novena":    "19°",
        "decimonoveno":     "19°", "decimonovena":     "19°",
        # ── 18 ──
        "décimo octavo":    "18°", "decimo octavo":    "18°",
        "décima octava":    "18°", "decima octava":    "18°",
        "decimoctavo":      "18°", "decimoctava":      "18°",
        # ── 17 ──
        "décimo séptimo":   "17°", "decimo septimo":   "17°",
        "décima séptima":   "17°", "decima septima":   "17°",
        "decimoséptimo":    "17°", "decimoseptimo":    "17°",
        "decimoséptima":    "17°", "decimoseptima":    "17°",
        # ── 16 ──
        "décimo sexto":     "16°", "decimo sexto":     "16°",
        "décima sexta":     "16°", "decima sexta":     "16°",
        "decimosexto":      "16°", "decimosexta":      "16°",
        # ── 15 ──
        "décimo quinto":    "15°", "decimo quinto":    "15°",
        "décima quinta":    "15°", "decima quinta":    "15°",
        "decimoquinto":     "15°", "decimoquinta":     "15°",
        # ── 14 ──
        "décimo cuarto":    "14°", "decimo cuarto":    "14°",
        "décima cuarta":    "14°", "decima cuarta":    "14°",
        "decimocuarto":     "14°", "decimocuarta":     "14°",
        # ── 13 ──
        "décimo tercero":   "13°", "decimo tercero":   "13°",
        "décimo tercer":    "13°", "decimo tercer":    "13°",
        "décima tercera":   "13°", "decima tercera":   "13°",
        "decimotercero":    "13°", "decimotercer":     "13°",
        "decimotercera":    "13°",
        # ── 12 ──
        "décimo segundo":   "12°", "decimo segundo":   "12°",
        "décima segunda":   "12°", "decima segunda":   "12°",
        "duodécimo":        "12°", "duodecimo":        "12°",
        "duodécima":        "12°", "duodecima":        "12°",
        "decimosegundo":    "12°", "decimosegunda":    "12°",
        # ── 11 ──
        "décimo primero":   "11°", "decimo primero":   "11°",
        "décimo primer":    "11°", "decimo primer":    "11°",
        "décima primera":   "11°", "decima primera":   "11°",
        "undécimo":         "11°", "undecimo":         "11°",
        "undécima":         "11°", "undecima":         "11°",
        "decimoprimero":    "11°", "decimoprimer":     "11°",
        "decimoprimera":    "11°",
        # ── 10 ──
        "décimo": "10°", "decimo": "10°",
        "décima": "10°", "decima": "10°",
        # ── 9 ──
        "noveno": "9°", "novena": "9°",
        # ── 8 ──
        "octavo": "8°", "octava": "8°",
        # ── 7 ──
        "séptimo": "7°", "septimo": "7°",
        "séptima": "7°", "septima": "7°",
        # ── 6 ──
        "sexto": "6°", "sexta": "6°",
        # ── 5 ──
        "quinto": "5°", "quinta": "5°",
        # ── 4 ──
        "cuarto": "4°", "cuarta": "4°",
        # ── 3 ──
        "tercero": "3°", "tercer": "3°",
        "tercera": "3°",
        # ── 2 ──
        "segundo": "2°", "segunda": "2°",
        # ── 1 ──
        "primero": "1°", "primer": "1°",
        "primera": "1°",
    }

    # Recorrer de mayor a menor longitud de clave para evitar colisiones de
    # subcadenas (ej: "décimo tercer" debe procesarse antes que "tercer").
    resultado = nombre
    for escrito in sorted(ORDINALES_MAP.keys(), key=len, reverse=True):
        patron = r'\b' + re.escape(escrito) + r'\b'
        resultado, n = re.subn(patron, ORDINALES_MAP[escrito], resultado, flags=re.IGNORECASE)
        if n > 0:
            break

    return resultado


def _extraer_ordinal(texto: str) -> int | None:
    """
    Extrae el número ordinal de un nombre de tribunal.
    Ej: "1° Juzgado Civil de Santiago" → 1
    Ej: "Decimocuarto Juzgado Civil" → 14
    """
    numerales = {
        "primer": 1, "primero": 1, "primera": 1,
        "segundo": 2, "segunda": 2,
        "tercero": 3, "tercera": 3, "tercer": 3,
        "cuarto": 4, "cuarta": 4,
        "quinto": 5, "quinta": 5,
        "sexto": 6, "sexta": 6,
        "séptimo": 7, "septimo": 7, "séptima": 7,
        "octavo": 8, "octava": 8,
        "noveno": 9, "novena": 9,
        "décimo": 10, "decimo": 10, "décima": 10,
        "decimoprimero": 11, "decimoprimer": 11, "undécimo": 11,
        "decimosegundo": 12, "duodécimo": 12,
        "decimotercero": 13, "decimotercer": 13,
        "decimocuarto": 14,
        "decimoquinto": 15,
        "decimosexto": 16,
        "decimoséptimo": 17, "decimoseptimo": 17,
        "decimoctavo": 18,
        "decimonoveno": 19,
        "vigésimo": 20, "vigesimo": 20,
    }
    texto_lower = texto.lower()
    # Número arábigo con símbolo ordinal
    m = re.search(r"(\d+)\s*[°ºª]", texto)
    if m:
        return int(m.group(1))
    # Número arábigo solo al inicio
    m = re.search(r"^(\d+)\s+", texto)
    if m:
        return int(m.group(1))
    # Numeral escrito
    for palabra, num in numerales.items():
        if palabra in texto_lower:
            return num
    return None


def _cargar_referencia_tribunales() -> list[dict[str, str]]:
    """
    Lee la hoja REFERENCIA de causas_ojv.xlsx.
    Retorna lista de dicts con claves: nombre_tribunal, corte.
    """
    try:
        wb = openpyxl.load_workbook(CAUSAS_XLSX, read_only=True, data_only=True)
    except FileNotFoundError:
        log.warning("causas_ojv.xlsx no encontrado en %s — buscar_corte deshabilitado", CAUSAS_XLSX)
        return []

    if "REFERENCIA" not in wb.sheetnames:
        log.warning("Hoja REFERENCIA no encontrada en %s", CAUSAS_XLSX)
        wb.close()
        return []

    ws = wb["REFERENCIA"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # Detectar encabezado: buscar columnas que contengan "tribunal" y "corte"
    header = [str(c).lower().strip() if c else "" for c in rows[0]]
    col_tribunal = next((i for i, h in enumerate(header) if "tribunal" in h), None)
    col_corte = next((i for i, h in enumerate(header) if "corte" in h), None)
    if col_tribunal is None or col_corte is None:
        log.warning(
            "Columnas tribunal/corte no encontradas en REFERENCIA (header=%s). "
            "Usando col_tribunal=%s, col_corte=%s",
            header, col_tribunal, col_corte,
        )
        col_tribunal = col_tribunal if col_tribunal is not None else 1
        col_corte = col_corte if col_corte is not None else 0
    log.debug("REFERENCIA columnas: tribunal=%d, corte=%d (header=%s)", col_tribunal, col_corte, header)

    resultado = []
    for fila in rows[1:]:
        nombre = fila[col_tribunal] if len(fila) > col_tribunal else None
        corte = fila[col_corte] if len(fila) > col_corte else None
        if nombre and corte:
            resultado.append({
                "nombre_tribunal": str(nombre).strip(),
                "corte": str(corte).strip(),
            })
    log.debug("REFERENCIA cargada: %d tribunales", len(resultado))
    return resultado


def _cargar_causas_historico() -> set[str]:
    """
    Lee la hoja CAUSAS de causas_ojv.xlsx.
    Retorna set de ROLes ya procesados (formato "ROL-AÑO").
    """
    try:
        wb = openpyxl.load_workbook(CAUSAS_XLSX, read_only=True, data_only=True)
    except FileNotFoundError:
        return set()

    if "CAUSAS" not in wb.sheetnames:
        wb.close()
        return set()

    ws = wb["CAUSAS"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return set()

    header = [str(c).lower().strip() if c else "" for c in rows[0]]
    col_rol = col_anio = None
    for i, h in enumerate(header):
        if "rol" in h and col_rol is None:
            col_rol = i
        if ("año" in h or "anio" in h or "year" in h) and col_anio is None:
            col_anio = i

    if col_rol is None:
        col_rol = 0
    if col_anio is None:
        col_anio = 1

    historico: set[str] = set()
    for fila in rows[1:]:
        rol = fila[col_rol] if len(fila) > col_rol else None
        anio = fila[col_anio] if len(fila) > col_anio else None
        if rol:
            key = f"{str(rol).strip()}-{str(anio).strip()}" if anio else str(rol).strip()
            historico.add(key)

    log.debug("Histórico CAUSAS: %d entradas", len(historico))
    return historico


_referencia_cache: list[dict[str, str]] | None = None


def buscar_corte(nombre_tribunal: str) -> str | None:
    """
    Busca la corte de apelaciones correspondiente a un tribunal usando RapidFuzz.
    Umbral: token_set_ratio >= 80, con validación ordinal post-matching.
    Retorna nombre de corte, o None si no se encuentra.
    """
    global _referencia_cache
    if _referencia_cache is None:
        _referencia_cache = _cargar_referencia_tribunales()

    if not _referencia_cache or not nombre_tribunal:
        return None

    nombre_limpio = _limpiar_tribunal(nombre_tribunal) or nombre_tribunal
    ordinal_query = _extraer_ordinal(nombre_limpio)

    mejor_score = 0
    mejor_corte = None
    mejor_tribunal = None

    for entry in _referencia_cache:
        score = fuzz.token_set_ratio(nombre_limpio.lower(), entry["nombre_tribunal"].lower())
        if score > mejor_score:
            mejor_score = score
            mejor_corte = entry["corte"]
            mejor_tribunal = entry["nombre_tribunal"]

    fuzzy_ok = True
    if mejor_score < _UMBRAL_FUZZY_TRIBUNAL:
        log.debug("Tribunal no encontrado (score=%d): %s", mejor_score, nombre_limpio)
        fuzzy_ok = False

    # Validación ordinal: si ambos tienen ordinal, deben coincidir
    if fuzzy_ok and ordinal_query is not None and mejor_tribunal is not None:
        ordinal_match = _extraer_ordinal(mejor_tribunal)
        if ordinal_match is not None and ordinal_query != ordinal_match:
            log.debug(
                "Ordinal mismatch (query=%d, match=%d) para: %s",
                ordinal_query, ordinal_match, nombre_limpio,
            )
            fuzzy_ok = False

    if fuzzy_ok:
        log.debug("Tribunal '%s' → corte '%s' (score=%d)", nombre_limpio, mejor_corte, mejor_score)
        return mejor_corte

    # Fallback: asignación directa por nombre de localidad
    nombre_lower = nombre_limpio.lower()
    _SAN_MIGUEL_KEYWORDS = (
        "san miguel", "san bernardo", "puente alto", "buin",
        "talagante", "colina", "melipilla", "peñaflor",
    )
    for kw in _SAN_MIGUEL_KEYWORDS:
        if kw in nombre_lower:
            log.debug("Fallback corte por nombre: '%s' → 'C.A. de San Miguel'", nombre_limpio)
            return "C.A. de San Miguel"
    if "santiago" in nombre_lower:
        log.debug("Fallback corte por nombre: '%s' → 'C.A. de Santiago'", nombre_limpio)
        return "C.A. de Santiago"

    log.debug("Tribunal sin corte (fuzzy ni fallback): %s", nombre_limpio)
    return None


# ===========================================================================
# LÓGICA DE PLAYWRIGHT
# ===========================================================================

def _construir_url_cuerpo_a(fecha: date) -> str:
    return f"{MERCURIO_BASE_URL}/{fecha.year}/{fecha.month:02d}/{fecha.day:02d}/A"


async def _esta_logueado(page: Page) -> bool:
    """Verifica si ya hay una sesión activa (el botón de login no es visible)."""
    try:
        btn = page.locator("#openPram")
        visible = await btn.is_visible()
        if not visible:
            return True
        # También puede estar visible pero con texto distinto post-login
        texto = (await btn.inner_text()).strip()
        return "iniciar" not in texto.lower()
    except Exception:
        return False


async def _hacer_login(page: Page) -> None:
    """Realiza el flujo de login con las credenciales de config.py."""
    log.info("Iniciando login en El Mercurio Digital…")

    # Abrir modal de login
    await page.locator("#openPram > span").click()
    await page.wait_for_timeout(1000)

    # Rellenar usuario
    await page.locator("#txtUsername").fill(MERCURIO_USER)
    await page.wait_for_timeout(300)

    # Rellenar contraseña
    await page.locator("#txtPassword").fill(MERCURIO_PASS)
    await page.wait_for_timeout(300)

    # Click en "Ingrese acá"
    async with page.expect_navigation(timeout=30_000):
        await page.locator("#gopram").click()

    await page.wait_for_timeout(1500)

    # Secuencia post-login completa (Scraper_Mercurio.json):
    # Escape ×2 → click fuera de #modal_mer_promoLS → (click CLASIFICADOS viene después)
    log.debug("Cerrando modales post-login: Escape ×2")
    await page.keyboard.press("Escape")
    await page.wait_for_timeout(300)
    await page.keyboard.press("Escape")
    await page.wait_for_timeout(500)

    # Click fuera de #modal_mer_promoLS (click en el overlay, fuera del contenido)
    try:
        promo = page.locator("#modal_mer_promoLS")
        if await promo.is_visible(timeout=3000):
            # Click en la esquina derecha del overlay (fuera del contenido del modal)
            box = await promo.bounding_box()
            if box:
                await page.mouse.click(box["x"] + box["width"] - 10, box["y"] + 10)
                log.debug("Click fuera de #modal_mer_promoLS para cerrarlo")
            else:
                await page.keyboard.press("Escape")
            await page.wait_for_timeout(500)
    except Exception:
        pass

    # Verificar login exitoso
    if not await _esta_logueado(page):
        raise RuntimeError(
            "Login fallido: el botón de login sigue visible. "
            "Verifica MERCURIO_USER y MERCURIO_PASS en config.py."
        )
    log.info("Login exitoso.")


async def _cerrar_modales(page: Page) -> None:
    """
    Cierra modales que puedan aparecer al navegar por la edición.
    Usa primero un cierre genérico vía jQuery (cubre modales futuros),
    luego cierra los conocidos por ID como fallback.
    """
    # --- Cierre genérico: ocultar TODOS los modales Bootstrap visibles ---
    try:
        await page.evaluate("""() => {
            if (typeof $ !== 'undefined') {
                $('.modal.in, .modal.show').modal('hide');
            }
        }""")
        await page.wait_for_timeout(500)
    except Exception:
        pass

    # Escape ×2 (cierra modales genéricos que jQuery no alcance)
    await page.keyboard.press("Escape")
    await page.wait_for_timeout(300)
    await page.keyboard.press("Escape")
    await page.wait_for_timeout(300)

    # Modales específicos de El Mercurio — click fuera (en overlay) como fallback
    for modal_id in ["#modal_mer_promoLS", "#modal_mer_promoINV", "#modal_mer_selectHome"]:
        try:
            modal = page.locator(modal_id)
            if await modal.is_visible(timeout=2000):
                box = await modal.bounding_box()
                if box:
                    await page.mouse.click(box["x"] + box["width"] - 10, box["y"] + 10)
                    log.debug("Cerrado modal %s (click fuera)", modal_id)
                else:
                    await page.keyboard.press("Escape")
                await page.wait_for_timeout(500)
        except Exception:
            pass

    # Fallback: cerrar cualquier modal Bootstrap restante vía botón .close
    for selector in [".modal.in .close", ".modal.show .close"]:
        try:
            elem = page.locator(selector).first
            if await elem.is_visible(timeout=1000):
                await elem.click()
                await page.wait_for_timeout(500)
        except Exception:
            pass


async def _verificar_fecha_edicion(page: Page, fecha_pedida: date) -> bool:
    """
    Compara la variable JS fechaEdicion de la página con la fecha solicitada.
    Retorna True si coinciden, False si no.
    """
    try:
        fecha_real = await page.evaluate("fechaEdicion")  # "2026/03/15"
    except Exception:
        fecha_real = None
    fecha_pedida_str = f"{fecha_pedida.year}/{fecha_pedida.month:02d}/{fecha_pedida.day:02d}"
    log.info("Fecha solicitada: %s, fecha edición cargada: %s", fecha_pedida_str, fecha_real)
    if fecha_real and fecha_real.strip() == fecha_pedida_str:
        return True
    return False


async def _navegar_a_sección_f(page: Page, fecha: date) -> bool:
    """
    Desde cuerpo A, intenta navegar a la sección F (Clasificados).
    Retorna True si la navegación fue exitosa, False si falló (modal bloqueante, etc.).
    """
    log.info("Navegando a sección F (Clasificados)…")

    # Cerrar modales ANTES de intentar click (previene bloqueo por promoINV, etc.)
    await _cerrar_modales(page)
    await page.wait_for_timeout(500)

    # Hacer clic en botón CLASIFICADOS del header
    clasificados_btn = page.locator("#uctHeader_ctl02_rptBodyPart_ctl07_aBody")
    try:
        await clasificados_btn.wait_for(state="visible", timeout=15_000)
    except Exception:
        clasificados_btn = page.locator("text=CLASIFICADOS")

    try:
        async with page.expect_navigation(
            url=lambda u: "/F" in u or "/f" in u,
            timeout=15_000,
        ):
            await clasificados_btn.click(timeout=15_000)

        await page.wait_for_timeout(1500)
        await _cerrar_modales(page)
        log.debug("Sección F cargada: %s", page.url)
        return True

    except Exception as e:
        log.warning("No se pudo navegar a sección F vía botón Clasificados: %s", e)
        return False


async def _navegar_directo_a_seccion(page: Page, fecha: date, seccion: str) -> None:
    """Navega directamente a una sección por URL (sin usar botón del header)."""
    url = f"{MERCURIO_BASE_URL}/{fecha.year}/{fecha.month:02d}/{fecha.day:02d}/{seccion}"
    log.info("Navegando directo a sección %s: %s", seccion, url)
    await page.goto(url, wait_until="domcontentloaded", timeout=30_000)
    await page.wait_for_timeout(2000)
    await _cerrar_modales(page)


async def _obtener_ids_paginas(page: Page, seccion: str) -> list[str]:
    """
    Extrae la lista ordenada de IDs de página de la sección indicada desde el DOM.
    seccion: 'F', 'B' o 'D' (o cualquier letra de cuerpo).
    Retorna una lista de strings con los IDs en orden de página.
    """
    ids = await page.evaluate("""
    (sec) => {
        // Buscar todos los enlaces con onclick="gotoPage('SEC', 'ID', NUM)"
        const pattern = new RegExp("gotoPage\\\\s*\\\\(\\\\s*'" + sec + "'\\\\s*,\\\\s*'([^']+)'\\\\s*,\\\\s*(\\\\d+)\\\\s*\\\\)");
        const seen = new Map();
        const allElems = document.querySelectorAll('[onclick*="gotoPage"]');
        for (const el of allElems) {
            const oc = el.getAttribute('onclick') || '';
            const m = pattern.exec(oc);
            if (m) {
                const pageId = m[1];
                const pageNum = parseInt(m[2], 10);
                if (!seen.has(pageId)) {
                    seen.set(pageId, pageNum);
                }
            }
        }
        // Convertir a array y ordenar por número de página
        const arr = Array.from(seen.entries())
                         .map(([id, num]) => ({ id, num }))
                         .sort((a, b) => a.num - b.num)
                         .map(x => x.id);
        return arr;
    }
    """, seccion)
    log.debug("IDs de páginas %s encontrados: %s", seccion, ids)
    return ids or []


async def _navegar_a_pagina(page: Page, fecha: date, page_id: str, seccion: str = "F") -> None:
    """Navega directamente al visor de una página específica del cuerpo indicado."""
    url = (
        f"{MERCURIO_BASE_URL}/{fecha.year}/{fecha.month:02d}/{fecha.day:02d}"
        f"/{seccion}/{page_id}#zoom=page-width"
    )
    log.debug("Navegando a página %s/%s  →  %s", seccion, page_id, url)
    await page.goto(url, wait_until="domcontentloaded", timeout=20_000)
    await page.wait_for_timeout(2000)
    await _cerrar_modales(page)


async def _leer_texto_layer(page: Page, max_wait_ms: int = 10_000) -> str:
    """
    Lee el contenido de texto del .textLayer de la página actual.
    Espera hasta max_wait_ms a que el textLayer tenga contenido.
    """
    inicio = time.time()
    while (time.time() - inicio) < (max_wait_ms / 1000):
        try:
            texto = await page.evaluate("""
            () => {
                const tl = document.querySelector('.textLayer');
                return tl ? tl.innerText : '';
            }
            """)
            if texto and texto.strip():
                return texto
        except Exception as e:
            log.debug("Error leyendo textLayer: %s", e)
        await page.wait_for_timeout(500)
    log.debug("textLayer vacío tras esperar %d ms", max_wait_ms)
    return ""


def _detectar_secciones(texto: str) -> list[str]:
    """Detecta las secciones numéricas presentes en el textLayer."""
    buscar = ["1611", "1612", "1613", "1614", "1615", "1616"]
    return [s for s in buscar if s in texto]


def _detectar_redireccion(texto: str) -> tuple[str, int] | None:
    """
    Busca en el textLayer un recuadro de redirección a otra sección.
    Ej: "MÁS AVISOS ECONÓMICOS CLASIFICADOS EN PÁG. C 8"
    Retorna (seccion, pagina) o None si no se encuentra.
    """
    m = _REDIRECT_PATTERN.search(texto)
    if m:
        seccion = m.group(1).upper()
        pagina = int(m.group(2))
        log.info("Redirección detectada: más avisos en sección %s, página %d", seccion, pagina)
        return (seccion, pagina)
    return None


async def _esperar_canvas_base(page: Page, timeout_ms: int = 15_000) -> bool:
    """Espera a que el canvas exista y tenga width > 0 (renderizado base)."""
    try:
        await page.wait_for_function(
            "document.querySelector('canvas#page1')?.width > 0"
            " || document.querySelector('#viewer canvas')?.width > 0",
            timeout=timeout_ms,
        )
        ancho = await page.evaluate("""
        () => {
            const c = document.querySelector('canvas#page1') ||
                      document.querySelector('#viewer canvas');
            return c ? c.width : 0;
        }
        """)
        log.debug("Canvas base renderizado: width=%d", ancho)
        return True
    except Exception as e:
        log.warning("Timeout esperando canvas base (width>0): %s", e)
        return False


async def _click_hd_btn(page: Page) -> bool:
    """Intenta clickear el botón HD. Retorna True si se hizo click."""
    try:
        hd_btn = page.locator("div.toolbar div.cont_activar_pdf > span:nth-of-type(1) img").first
        if await hd_btn.is_visible(timeout=5000):
            await hd_btn.click()
            log.debug("Botón HD clickeado (selector toolbar).")
            return True
        hd_btn2 = page.locator("#inactive_pdf img").first
        if await hd_btn2.is_visible(timeout=3000):
            await hd_btn2.click()
            log.debug("Botón HD clickeado (fallback #inactive_pdf).")
            return True
        log.warning("Botón HD no visible con ningún selector.")
        return False
    except Exception as e:
        log.warning("No se pudo clickear botón HD: %s", e)
        return False


async def _activar_hd(page: Page) -> None:
    """
    Activa el modo HD del visor:
    1. Espera a que canvas base renderice (width > 0)
    2. Clickea botón HD
    3. Si canvas sigue en 0 tras 5s, reintenta click
    """
    # 1. Esperar canvas base
    log.debug("Esperando a que canvas base renderice (width > 0)…")
    canvas_ok = await _esperar_canvas_base(page)
    if not canvas_ok:
        log.warning("Canvas base no renderizó; intentando HD de todas formas.")

    # 2. Primer click HD
    clicked = await _click_hd_btn(page)
    if not clicked:
        return

    # 3. Verificar si canvas reacciona; si sigue en 0, reintentar
    await page.wait_for_timeout(5000)
    try:
        ancho = await page.evaluate("""
        () => {
            const c = document.querySelector('canvas#page1') ||
                      document.querySelector('#viewer canvas');
            return c ? c.width : 0;
        }
        """)
        if ancho == 0:
            log.warning("Canvas sigue en width=0 tras primer click HD; reintentando click…")
            await _click_hd_btn(page)
        else:
            log.debug("Canvas post-HD click: width=%d", ancho)
    except Exception:
        pass


async def _esperar_canvas_hd(page: Page, timeout_ms: int = 20_000) -> bool:
    """
    Espera a que el canvas renderice en HD (width > 1800).
    Loguea el estado del canvas cada 2 segundos.
    Retorna True si se alcanzó HD, False si se agotó el timeout.
    """
    inicio = time.time()
    timeout_s = timeout_ms / 1000
    ultimo_log = 0.0

    while True:
        elapsed = time.time() - inicio
        if elapsed >= timeout_s:
            break

        try:
            ancho = await page.evaluate("""
            () => {
                const canvas = document.querySelector('canvas#page1') ||
                               document.querySelector('#viewer canvas');
                return canvas ? canvas.width : 0;
            }
            """)
            if ancho and int(ancho) > _CANVAS_HD_UMBRAL:
                log.debug("Canvas HD detectado: width=%d (%.0fs/%.0fs)", ancho, elapsed, timeout_s)
                return True

            # Log cada 2 segundos
            if elapsed - ultimo_log >= 2.0:
                log.debug("Esperando HD: canvas.width=%d (%.0fs/%.0fs)", ancho or 0, elapsed, timeout_s)
                ultimo_log = elapsed
        except Exception:
            pass
        await page.wait_for_timeout(500)

    # Log final
    try:
        ancho_final = await page.evaluate("""
        () => {
            const canvas = document.querySelector('canvas#page1') ||
                           document.querySelector('#viewer canvas');
            return canvas ? canvas.width : 0;
        }
        """)
        log.warning(
            "Timeout esperando HD (canvas.width=%d, umbral=%d). "
            "Continuando con resolución disponible.",
            ancho_final, _CANVAS_HD_UMBRAL,
        )
    except Exception:
        pass
    return False


# ---------------------------------------------------------------------------
# Troceo de páginas grandes (mitigación truncamiento DeepSeek)
# ---------------------------------------------------------------------------

def _trocear_pagina(texto: str, solape: int = 2000) -> list[str]:
    """Divide texto grande en 2 mitades con solape para no partir avisos a la mitad.

    - Si len(texto) <= 45000: devuelve [texto] (una sola pieza).
    - Si es mayor: busca el separador de párrafo más cercano al punto medio
      (primero "\\n\\n", luego "\\n", luego " "), corta ahí, y devuelve dos
      chunks con 'solape' caracteres de superposición. Así un aviso partido
      en el borde aparece íntegro en al menos uno de los dos chunks.
    """
    UMBRAL = 45_000
    if len(texto) <= UMBRAL:
        return [texto]

    medio = len(texto) // 2
    # Ventana de búsqueda: ±3000 chars alrededor del medio
    ventana_inicio = max(0, medio - 3_000)
    ventana_fin = min(len(texto), medio + 3_000)

    for sep in ("\n\n", "\n", " "):
        pos = texto.rfind(sep, ventana_inicio, ventana_fin)
        if pos != -1:
            corte = pos + len(sep)
            break
    else:
        corte = medio

    inicio_b = max(0, corte - solape)
    return [texto[:corte], texto[inicio_b:]]


# ===========================================================================
# TEXT API (Claude)
# ===========================================================================

def _enviar_texto_a_claude(page_id: str, texto: str, reintentos: int = 2) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """
    Envía texto del textLayer a la API de extracción (Sonnet o DeepSeek).
    Retorna (avisos, usage_dict) donde usage_dict = {"in": prompt_tokens, "out": completion_tokens}.
    Reintenta hasta 2 veces en caso de fallo (3 intentos totales).
    En fallo total retorna ([], {"in":0,"out":0}).
    """
    log.info("Motor de extracción: %s", MODELO_EXTRACCION)

    log.info(
        "Enviando texto pág %s a la API (%d caracteres)",
        page_id, len(texto),
    )

    contenido = PROMPT_EXTRACCION + "\n\n---\nTEXTO DE LA PÁGINA:\n" + texto

    for intento in range(reintentos + 1):
        try:
            log.info(
                "API pág %s (intento %d/%d, motor=%s)",
                page_id, intento + 1, reintentos + 1, MODELO_EXTRACCION,
            )

            if MODELO_EXTRACCION == "deepseek":
                # ── Rama DeepSeek V4-Flash (OpenAI SDK) ──
                from openai import OpenAI

                client = OpenAI(
                    api_key=DEEPSEEK_API_KEY,
                    base_url="https://api.deepseek.com",
                    timeout=60.0,
                )
                response = client.chat.completions.create(
                    model="deepseek-v4-flash",
                    max_tokens=16384,
                    messages=[{"role": "user", "content": contenido}],
                )
                raw = response.choices[0].message.content
                if not raw or not raw.strip():
                    raise ValueError("API devolvió respuesta vacía (posible cuelgue de pasarela)")
                texto_respuesta = raw
                usage = {"in": 0, "out": 0}
                if response.usage:
                    usage["in"] = response.usage.prompt_tokens or 0
                    usage["out"] = response.usage.completion_tokens or 0
            else:
                # ── Rama Sonnet (Anthropic SDK, actual) ──
                client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=60.0)
                response = client.messages.create(
                    model="claude-sonnet-4-6",
                    max_tokens=16384,
                    messages=[{
                        "role": "user",
                        "content": contenido,
                    }],
                )
                texto_respuesta = "".join(
                    bloque.text for bloque in response.content if hasattr(bloque, "text")
                )
                usage = {"in": response.usage.input_tokens or 0,
                        "out": response.usage.output_tokens or 0}

            log.debug("Respuesta API (primeros 200 chars): %s", texto_respuesta[:200])
            avisos = _parsear_json_vision(texto_respuesta)
            if not avisos and texto_respuesta.strip() not in ("[]", "[ ]"):
                raise ValueError(
                    f"Respuesta vacía o JSON truncado/inválido "
                    f"({len(texto_respuesta)} chars crudos)"
                )
            return (avisos, usage)

        except Exception as e:
            log.warning("Fallo en API/parseo pág %s (intento %d/%d): %s",
                        page_id, intento + 1, reintentos + 1, e)
            if intento < reintentos:
                time.sleep(5)

    log.error("API falló tras %d intentos para pág %s (motor=%s)", reintentos + 1, page_id, MODELO_EXTRACCION)
    return ([], {"in": 0, "out": 0})


def _parsear_json_vision(texto: str) -> list[dict[str, Any]]:
    """
    Parsea la respuesta de Vision API.
    Usa raw_decode para tolerar texto extra antes/después del JSON.
    """
    texto = texto.strip()
    # Quitar bloques de código markdown si los hay
    texto = re.sub(r"^```(?:json)?\s*", "", texto, flags=re.MULTILINE)
    texto = re.sub(r"```\s*$", "", texto, flags=re.MULTILINE)
    texto = texto.strip()

    decoder = json.JSONDecoder()
    # Buscar primer '[' para arrays
    idx = texto.find("[")
    if idx == -1:
        log.warning("Respuesta Vision no contiene array JSON: %s", texto[:200])
        return []
    try:
        resultado, _ = decoder.raw_decode(texto, idx)
        if isinstance(resultado, list):
            return resultado
    except json.JSONDecodeError as e:
        log.warning("JSON inválido en respuesta Vision: %s — texto: %s", e, texto[:300])

    return []


# ===========================================================================
# POST-PROCESAMIENTO
# ===========================================================================

_CBR_ANIO_MIN = 1900
_CBR_ANIO_MAX = 2027

def _evaluar_cbr_por_anio(anio_raw) -> dict:
    """
    Decide CBR a partir del año de inscripción del dominio que devolvió la IA.
    No usa regex sobre texto; opera sobre el año ya extraído.
    """
    if anio_raw is None or str(anio_raw).strip() == "":
        return {"decision": "REVISAR", "cbr_anio": None,
                "cbr_flag_revision": True, "cbr_motivo": "Año CBR no detectado"}
    m = re.search(r"\b(\d{4})\b", str(anio_raw))
    if not m:
        return {"decision": "REVISAR", "cbr_anio": None,
                "cbr_flag_revision": True, "cbr_motivo": "Año CBR no detectado"}
    anio = int(m.group(1))
    if not (_CBR_ANIO_MIN <= anio <= _CBR_ANIO_MAX):
        return {"decision": "REVISAR", "cbr_anio": None,
                "cbr_flag_revision": True, "cbr_motivo": "Año CBR fuera de rango"}
    if anio >= _CBR_ANIO_CORTE:
        return {"decision": "EXCLUIR", "cbr_anio": anio,
                "cbr_flag_revision": False, "cbr_motivo": ""}
    return {"decision": "MANTENER", "cbr_anio": anio,
            "cbr_flag_revision": False, "cbr_motivo": ""}

def _normalizar_aviso(raw: dict[str, Any]) -> dict[str, Any] | None:
    """
    Normaliza y valida un aviso crudo de Vision API.
    Retorna dict con el contrato de datos, o None si el aviso es inválido.
    """
    rol_raw = str(raw.get("rol") or "").strip()
    año = str(raw.get("año") or "").strip()
    tribunal_raw = raw.get("tribunal") or raw.get("juzgado")
    demandante = str(raw.get("demandante") or "").strip() or None
    demandado = str(raw.get("demandado") or "").strip() or None
    direccion = str(raw.get("direccion") or "").strip() or None
    comuna = str(raw.get("comuna") or "").strip() or None

    # Extraer año del ROL si viene en formato C-XXXXX-YYYY o XXXXX-YYYY
    rol = rol_raw.lstrip("Cc-").strip()
    if not año:
        # Intentar separar "12345-2024" → rol=12345, año=2024
        m = re.match(r"^(\d+)-(\d{4})$", rol)
        if m:
            rol, año = m.group(1), m.group(2)
            log.debug("Año extraído del ROL: %s → rol=%s, año=%s", rol_raw, rol, año)
        else:
            # Intentar desde rol_raw: "C-12345-2024"
            m = re.match(r"^[Cc]-?(\d+)-(\d{4})$", rol_raw)
            if m:
                rol, año = m.group(1), m.group(2)
                log.debug("Año extraído del ROL completo: %s → rol=%s, año=%s", rol_raw, rol, año)

    # Validaciones mínimas: ROL es obligatorio
    if not rol:
        log.debug("Aviso descartado por falta de ROL: %s", raw)
        return None
    if not re.match(r"^\d+$", rol):
        log.debug("ROL no numérico, descartando: %s", rol)
        return None

    # Si año sigue vacío, advertir pero dejar pasar para que M2 intente completar
    if not año:
        log.warning("Aviso sin AÑO (se envía a M2 para completar): rol=%s, raw=%s", rol, raw)

    # Limpiar nombre de tribunal
    tribunal_limpio = _limpiar_tribunal(str(tribunal_raw).strip() if tribunal_raw else None)
    tribunal_limpio = _normalizar_ordinal_tribunal(tribunal_limpio)

    # Mapear tribunal → corte
    corte = buscar_corte(tribunal_limpio) if tribunal_limpio else None

    return {
        "rol": rol,
        "año": año,
        "corte": corte or "",
        "tribunal": tribunal_limpio or "",
        "demandante": demandante or "",
        "demandado": demandado or "",
        "direccion": direccion,
        "comuna": comuna,
        "año_inscripcion_dominio": raw.get("año_inscripcion_dominio"),
        "fojas": raw.get("fojas"),
        "region_rm": True,
    }


def _filtrar_avisos(
    avisos: list[dict[str, Any]],
    historico: set[str],
    vistos_en_ejecucion: set[str],
    st=None,
) -> list[dict[str, Any]]:
    """
    Aplica todos los filtros del negocio a la lista de avisos normalizados.
    Modifica vistos_en_ejecucion in-place para deduplicar entre páginas.
    Loggea conteo antes/después por cada filtro.
    """
    total_entrada = len(avisos)
    desc_rm = desc_banco = desc_comuna = desc_anio = desc_hist = desc_dup = 0

    resultado = []
    for aviso in avisos:
        rol = aviso["rol"]
        año = aviso["año"]
        key = f"{rol}-{año}"

        # Filtro 1: Solo RM
        corte = aviso.get("corte", "")
        if corte not in _CORTES_RM:
            desc_rm += 1
            log.debug("  Descartado (no RM): ROL %s, corte='%s'", rol, corte)
            if st is not None:
                st.descartados.append({"rol": rol, "año": año, "tribunal": aviso.get("tribunal"), "motivo": "Solo RM", "monto": None})
            continue

        # Filtro 2: Banco Estado
        demandante_lower = (aviso.get("demandante") or "").lower()
        if any(b in demandante_lower for b in _BANCOS_ESTADO):
            desc_banco += 1
            log.debug("  Descartado (Banco Estado): ROL %s", rol)
            if st is not None:
                st.descartados.append({"rol": rol, "año": año, "tribunal": aviso.get("tribunal"), "motivo": "Banco Estado", "monto": None})
            continue

        # Filtro 3: Estación Central
        comuna_lower = (aviso.get("comuna") or "").lower().strip()
        if comuna_lower in _COMUNAS_EXCLUIDAS:
            desc_comuna += 1
            log.debug("  Descartado (Estación Central): ROL %s", rol)
            if st is not None:
                st.descartados.append({"rol": rol, "año": año, "tribunal": aviso.get("tribunal"), "motivo": "Estación Central", "monto": None})
            continue

        # Filtro 4: Año >= 2018  (renumerado)
        try:
            if int(año) < 2018:
                desc_anio += 1
                log.debug("  Descartado (pre-2018): ROL %s, año %s", rol, año)
                if st is not None:
                    st.descartados.append({"rol": rol, "año": año, "tribunal": aviso.get("tribunal"), "motivo": "Pre-2018", "monto": None})
                continue
        except ValueError:
            desc_anio += 1
            log.debug("  Año no parseable, descartando: %s", año)
            continue

        # Filtro 5: Dedup contra historial CAUSAS
        if key in historico:
            desc_hist += 1
            log.debug("  Descartado (ya en historial): ROL %s-%s", rol, año)
            continue

        # Filtro 6: Dedup entre páginas de la misma ejecución
        if key in vistos_en_ejecucion:
            desc_dup += 1
            log.debug("  Descartado (duplicado en ejecución): ROL %s-%s", rol, año)
            continue

        vistos_en_ejecucion.add(key)
        resultado.append(aviso)

    # Resumen de filtros
    log.info("Filtro Solo RM         : %d → %d (-%d)",
             total_entrada, total_entrada - desc_rm, desc_rm)
    post_rm = total_entrada - desc_rm
    log.info("Filtro Banco Estado    : %d → %d (-%d)",
             post_rm, post_rm - desc_banco, desc_banco)
    post_banco = post_rm - desc_banco
    log.info("Filtro Estación Central: %d → %d (-%d)",
             post_banco, post_banco - desc_comuna, desc_comuna)
    post_comuna = post_banco - desc_comuna
    log.info("Filtro Año >= 2018     : %d → %d (-%d)",
             post_comuna, post_comuna - desc_anio, desc_anio)
    post_anio = post_comuna - desc_anio
    log.info("Filtro Historial CAUSAS: %d → %d (-%d)",
             post_anio, post_anio - desc_hist, desc_hist)
    post_hist = post_anio - desc_hist
    log.info("Filtro Dup ejecución   : %d → %d (-%d)",
             post_hist, post_hist - desc_dup, desc_dup)
    log.info("Resultado final filtrado: %d de %d avisos pasan", len(resultado), total_entrada)

    return resultado


# ===========================================================================
# FUNCIÓN PRINCIPAL ASYNC
# ===========================================================================

async def _extraer_mercurio_async(
    fecha: date, *, dry_run: bool = False
) -> list[dict[str, Any]]:
    """
    Núcleo async del extractor. Abre Playwright, navega el diario, lee el
    textLayer de las páginas 1616 y las envía a Claude Text API.

    Si dry_run=True, ejecuta solo la navegación (login, sección F, detección de
    páginas 1616 y lectura del textLayer) pero NO envía nada a Claude API.
    """
    log_file = _setup_logging()
    st = _Stats()

    seccion_activa = "F"  # por defecto; puede cambiar a "B" (L-V) o "D" (fin de semana)

    log.info("=== Inicio extracción El Mercurio Digital ===")
    log.info("Fecha edición: %s | dry_run: %s", fecha.isoformat(), dry_run)
    log.info("Log file: %s", log_file)

    historico = _cargar_causas_historico()
    log.info("Histórico CAUSAS cargado: %d entradas", len(historico))
    vistos_en_ejecucion: set[str] = set()
    todas_las_causas: list[dict[str, Any]] = []
    paginas_texto: list[tuple[str, str]] = []  # [(page_id, texto_completo)]

    async with async_playwright() as pw:
        _base_dir = Path(os.path.dirname(os.path.abspath(__file__)))
        profile_dir = str(_base_dir / "playwright_profile")
        log.info("Lanzando Chromium headless (perfil: %s)", profile_dir)
        context = await pw.chromium.launch_persistent_context(
            user_data_dir=profile_dir,
            headless=True,
            viewport={"width": 1990, "height": 1279},
            java_script_enabled=True,
            accept_downloads=False,
        )
        page = context.pages[0] if context.pages else await context.new_page()

        try:
            # ---------------------------------------------------------------
            # Paso 1: Abrir Cuerpo A
            # ---------------------------------------------------------------
            url_a = _construir_url_cuerpo_a(fecha)
            log.info("[Paso 1/6] Navegando a cuerpo A: %s", url_a)
            await page.goto(url_a, wait_until="domcontentloaded", timeout=30_000)
            await page.wait_for_timeout(2000)
            log.info("[Paso 1/6] Cuerpo A cargado OK")

            # ---------------------------------------------------------------
            # Paso 2: Login si es necesario
            # ---------------------------------------------------------------
            log.info("[Paso 2/6] Verificando sesión…")
            if not await _esta_logueado(page):
                log.info("[Paso 2/6] Sesión no activa — iniciando login")
                await _hacer_login(page)
                log.info("[Paso 2/6] Login completado OK")
            else:
                log.info("[Paso 2/6] Sesión activa detectada, omitiendo login")

            # Siempre cerrar modales antes de navegar (login o no)
            await _cerrar_modales(page)

            # ---------------------------------------------------------------
            # Paso 3: Navegar a sección de clasificados
            #   - Intenta F vía botón (funciona L-V y domingos si F está actualizada)
            #   - Si F falla o tiene fecha stale:
            #       Sábado → sección D (Clasificados separados de B los sábados)
            #       L-V    → sección B (Clasificados al final de Economía y Negocios)
            #       Domingo→ sección D primero, luego B como último fallback
            # ---------------------------------------------------------------
            es_sabado = fecha.weekday() == 5   # 5 = sábado
            es_domingo = fecha.weekday() == 6  # 6 = domingo
            es_finde = es_sabado or es_domingo

            log.info(
                "[Paso 3/6] Navegando a clasificados (día=%s, finde=%s)",
                fecha.strftime("%A"), es_finde,
            )

            # --- Intento 1: Sección F vía botón Clasificados del header ---
            f_ok = await _navegar_a_sección_f(page, fecha)
            if f_ok:
                fecha_ok_f = await _verificar_fecha_edicion(page, fecha)
                if fecha_ok_f:
                    log.info("Sección F tiene la fecha correcta.")
                    seccion_activa = "F"
                else:
                    log.warning(
                        "Sección F no actualizada (no coincide con %s).",
                        fecha.isoformat(),
                    )
                    f_ok = False
            else:
                log.warning("Click a Clasificados falló. Intentando fallback…")

            # --- Fallback: si F no funcionó ---
            if not f_ok:
                if es_finde:
                    # Fines de semana: clasificados están en sección D
                    log.info("Fin de semana detectado → intentando sección D (Clasificados)")
                    await _navegar_directo_a_seccion(page, fecha, "D")
                    fecha_ok_d = await _verificar_fecha_edicion(page, fecha)
                    if fecha_ok_d:
                        seccion_activa = "D"
                        log.info("Sección D tiene la fecha correcta. Continuando con sección D.")
                    else:
                        # Último intento: sección B
                        log.warning(
                            "Sección D no tiene fecha %s. Intentando sección B…",
                            fecha.isoformat(),
                        )
                        await _navegar_directo_a_seccion(page, fecha, "B")
                        fecha_ok_b = await _verificar_fecha_edicion(page, fecha)
                        if fecha_ok_b:
                            seccion_activa = "B"
                            log.info("Sección B tiene la fecha correcta. Continuando con sección B.")
                        else:
                            log.error(
                                "Ni sección F, D ni B tienen la edición del %s. "
                                "El diario probablemente no ha sido publicado aún.",
                                fecha.isoformat(),
                            )
                            log.error(
                                "Ni seccion F ni D ni B tienen la edicion del %s.",
                                fecha.isoformat(),
                            )
                            _log_resumen(st, dry_run=dry_run)
                            raise EdicionNoDisponible(fecha.isoformat())
                else:
                    # Días de semana (L-V): clasificados al final de sección B
                    log.info("Día de semana → intentando sección B (Economía y Negocios)")
                    await _navegar_directo_a_seccion(page, fecha, "B")
                    fecha_ok_b = await _verificar_fecha_edicion(page, fecha)
                    if not fecha_ok_b:
                        log.error(
                            "Ni sección F ni B tienen la edición del %s. "
                            "El diario probablemente no ha sido publicado aún.",
                            fecha.isoformat(),
                        )
                        log.error(
                            "Ni seccion F ni B no tienen la edicion del %s.",
                            fecha.isoformat(),
                        )
                        _log_resumen(st, dry_run=dry_run)
                        raise EdicionNoDisponible(fecha.isoformat())
                    seccion_activa = "B"
                    log.info("Sección B tiene la fecha correcta. Continuando con sección B.")

            st.seccion_utilizada = seccion_activa

            # ---------------------------------------------------------------
            # Paso 4: Obtener lista de IDs de páginas
            # ---------------------------------------------------------------
            log.info("[Paso 4/6] Obteniendo mapa de páginas de sección %s", seccion_activa)
            ids_paginas = await _obtener_ids_paginas(page, seccion_activa)
            if len(ids_paginas) < 1:
                log.error(
                    "[Paso 4/6] Insuficientes IDs de páginas %s (encontrados: %d). "
                    "Posible error de carga o edición no disponible. Abortando.",
                    seccion_activa, len(ids_paginas),
                )
                _log_resumen(st, dry_run=dry_run)
                return []

            log.info("[Paso 4/6] Páginas %s encontradas: %d — inicio en última (índice %d)",
                     seccion_activa, len(ids_paginas), len(ids_paginas) - 1)
            indice_inicio = len(ids_paginas) - 1

            # ---------------------------------------------------------------
            # Paso 5: Navegar a última página y activar HD (una sola vez)
            # ---------------------------------------------------------------
            ultima_id = ids_paginas[indice_inicio]
            log.info("[Paso 5/6] Navegando a última página %s/%s para activar HD", seccion_activa, ultima_id)
            await _navegar_a_pagina(page, fecha, ultima_id, seccion_activa)

            log.info("Activando modo HD (una sola vez para toda la sesión)…")
            await _activar_hd(page)
            hd_ok = await _esperar_canvas_hd(page, timeout_ms=20_000)
            if hd_ok:
                log.info("Canvas HD confirmado (width > %d). HD queda activo para toda la sesión.", _CANVAS_HD_UMBRAL)
            else:
                log.warning("HD no confirmado, continuando con resolución disponible.")
            await page.wait_for_timeout(2000)  # buffer post-renderizado

            # ---------------------------------------------------------------
            # Paso 6: Loop retroceder desde última (tope 15 páginas)
            # ---------------------------------------------------------------
            log.info("[Paso 6/6] Iniciando recorrido hacia atrás (máx %d páginas)", _MAX_PAGINAS)
            indice_actual = indice_inicio

            while st.paginas_revisadas < _MAX_PAGINAS and indice_actual >= 0:
                page_id = ids_paginas[indice_actual]
                st.paginas_revisadas += 1
                log.info(
                    "--- Página %s (índice %d/%d, revisada #%d) ---",
                    page_id, indice_actual + 1, len(ids_paginas), st.paginas_revisadas,
                )

                # Navegar (salvo la primera iteración, ya estamos en última)
                if indice_actual != indice_inicio:
                    try:
                        await _navegar_a_pagina(page, fecha, page_id, seccion_activa)
                    except Exception as e:
                        log.warning("Error navegando a página %s: %s — saltando", page_id, e)
                        indice_actual -= 1
                        continue

                # Buffer de 2s para que el textLayer HD se estabilice
                await page.wait_for_timeout(2000)

                # Leer textLayer completo
                texto_layer = await _leer_texto_layer(page)
                log.debug(
                    "textLayer pág %s (300 chars): \"%s\"",
                    page_id, texto_layer[:300].replace("\n", "\\n"),
                )

                # Detectar secciones
                secciones = _detectar_secciones(texto_layer)
                log.debug("Secciones detectadas en pág %s: %s", page_id, secciones)

                contiene_1616 = "1616" in secciones
                tiene_menor = bool(set(secciones) & _SECCIONES_MENORES)

                # Decisión
                if not contiene_1616:
                    log.info(
                        "Pág %s: contiene 1616=No, sección menor=N/A → acción: descartar",
                        page_id,
                    )
                    st.paginas_descartadas += 1
                elif contiene_1616 and not tiene_menor:
                    # Contiene 1616 sin sección menor → conservar, continuar
                    log.info(
                        "Pág %s: contiene 1616=Sí, sección menor=No → acción: conservar",
                        page_id,
                    )
                    st.paginas_con_1616 += 1
                    paginas_texto.append((page_id, texto_layer))
                else:
                    # Contiene 1616 Y sección menor → conservar y PARAR
                    log.info(
                        "Pág %s: contiene 1616=Sí, sección menor=Sí (%s) → acción: PARAR (inicio de 1616)",
                        page_id, [s for s in secciones if s in _SECCIONES_MENORES],
                    )
                    st.paginas_con_1616 += 1
                    st.pagina_parada = page_id
                    paginas_texto.append((page_id, texto_layer))
                    break  # PARAR

                indice_actual -= 1

            # Log de condición de parada
            if st.pagina_parada:
                log.info("Parada: inicio de sección 1616 detectado en página %s", st.pagina_parada)
            elif st.paginas_revisadas >= _MAX_PAGINAS:
                log.warning("Tope de seguridad alcanzado: %d páginas revisadas", _MAX_PAGINAS)
            elif indice_actual < 0:
                log.warning("Se llegó al inicio de la sección %s sin encontrar inicio de 1616", seccion_activa)

            # ---------------------------------------------------------------
            # Paso 6b: Revisar cachito de 1616 en últimas 3 páginas de B
            # Siempre que la sección primaria NO sea B, puede haber avisos
            # 1616 sueltos al final de la sección B (Economía y Negocios).
            # HD persiste en la sesión, no necesita reactivarse.
            # ---------------------------------------------------------------
            if seccion_activa != "B":
                log.info("[Paso 6b] Revisando últimas páginas de sección B por avisos 1616 adicionales")
                try:
                    await _navegar_directo_a_seccion(page, fecha, "B")
                    ids_b = await _obtener_ids_paginas(page, "B")
                    if ids_b:
                        n_revisar = min(3, len(ids_b))
                        ultimas_b = ids_b[-n_revisar:]
                        log.info(
                            "Revisando %d últimas páginas de B: %s",
                            n_revisar, ultimas_b,
                        )
                        for page_id_b in reversed(ultimas_b):
                            try:
                                await _navegar_a_pagina(page, fecha, page_id_b, "B")
                                await page.wait_for_timeout(2000)
                                texto_b = await _leer_texto_layer(page)
                                secciones_b = _detectar_secciones(texto_b)
                                st.paginas_revisadas += 1

                                if "1616" in secciones_b:
                                    log.info(
                                        "Pág B/%s: contiene 1616 → conservar (cachito B)",
                                        page_id_b,
                                    )
                                    paginas_texto.append((page_id_b, texto_b))
                                    st.paginas_con_1616 += 1
                                else:
                                    log.debug(
                                        "Pág B/%s: sin 1616 → descartar",
                                        page_id_b,
                                    )
                                    st.paginas_descartadas += 1
                            except Exception as e:
                                log.warning(
                                    "Error revisando pág B/%s: %s — saltando",
                                    page_id_b, e,
                                )
                    else:
                        log.warning("No se encontraron páginas en sección B")
                except Exception as e:
                    log.warning(
                        "Error accediendo a sección B para cachito: %s — continuando sin cachito",
                        e,
                    )

            # ---------------------------------------------------------------
            # Paso 6c: Redirección a otra sección desde B
            # Alguna página conservada puede contener un recuadro:
            #   "MÁS AVISOS ECONÓMICOS CLASIFICADOS EN PÁG. C 8"
            # Buscamos en TODAS las páginas conservadas (el recuadro suele
            # estar en la última página con 1616, no en la última absoluta
            # de B que puede ser contenido editorial).
            # Si existe, navegar a esa sección/página y leer hacia adelante
            # mientras haya contenido 1616.
            # HD persiste en la sesión, no necesita reactivarse.
            # ---------------------------------------------------------------
            redir = None
            for _pid, _txt in paginas_texto:
                redir = _detectar_redireccion(_txt)
                if redir:
                    log.info(
                        "[Paso 6c] Redirección encontrada en página %s", _pid,
                    )
                    break

            if redir:
                seccion_redir, pagina_redir = redir
                log.info(
                    "[Paso 6c] Redirección detectada → sección %s, página %d",
                    seccion_redir, pagina_redir,
                )
                try:
                    await _navegar_directo_a_seccion(page, fecha, seccion_redir)
                    ids_redir = await _obtener_ids_paginas(page, seccion_redir)
                    if ids_redir:
                        log.info(
                            "Sección %s tiene %d páginas: %s",
                            seccion_redir, len(ids_redir), ids_redir,
                        )
                        # pagina_redir es número de página (1-based),
                        # ids_redir es lista ordenada por número
                        idx_inicio = pagina_redir - 1  # 0-based
                        if idx_inicio < 0 or idx_inicio >= len(ids_redir):
                            log.warning(
                                "Página %d fuera de rango en sección %s (%d páginas). "
                                "Intentando desde la última.",
                                pagina_redir, seccion_redir, len(ids_redir),
                            )
                            idx_inicio = len(ids_redir) - 1

                        # Leer HACIA ADELANTE desde la página indicada
                        idx_redir = idx_inicio
                        while idx_redir < len(ids_redir):
                            page_id_r = ids_redir[idx_redir]
                            try:
                                await _navegar_a_pagina(
                                    page, fecha, page_id_r, seccion_redir,
                                )
                                await page.wait_for_timeout(2000)
                                texto_r = await _leer_texto_layer(page)
                                secciones_r = _detectar_secciones(texto_r)
                                st.paginas_revisadas += 1

                                if "1616" in secciones_r:
                                    log.info(
                                        "Pág %s/%s (pág %d): contiene 1616 → conservar (cachito %s)",
                                        seccion_redir, page_id_r,
                                        idx_redir + 1, seccion_redir,
                                    )
                                    paginas_texto.append((page_id_r, texto_r))
                                    st.paginas_con_1616 += 1
                                    idx_redir += 1
                                else:
                                    log.info(
                                        "Pág %s/%s (pág %d): sin 1616 → parar lectura %s",
                                        seccion_redir, page_id_r,
                                        idx_redir + 1, seccion_redir,
                                    )
                                    st.paginas_descartadas += 1
                                    break
                            except Exception as e:
                                log.warning(
                                    "Error leyendo pág %s/%s: %s — saltando",
                                    seccion_redir, page_id_r, e,
                                )
                                idx_redir += 1
                    else:
                        log.warning(
                            "No se encontraron páginas en sección %s",
                            seccion_redir,
                        )
                except Exception as e:
                    log.warning(
                        "Error accediendo a sección %s para cachito: %s — continuando",
                        seccion_redir, e,
                    )
            else:
                log.debug("[Paso 6c] Sin redirección detectada en páginas conservadas")

        except Exception as e:
            log.error("Error crítico durante la navegación: %s", e, exc_info=True)
        finally:
            await context.close()
            log.info("Navegador cerrado")

    # -----------------------------------------------------------------------
    # Paso 7 & 8: Enviar texto a Claude API y filtrar (saltar si dry_run)
    # -----------------------------------------------------------------------
    if dry_run:
        log.info("[Paso 7/8] OMITIDO (dry run) — Claude API no invocada")
        log.info("[Paso 8/8] OMITIDO (dry run) — Filtrado y dedup no aplicados")
        log.info("DRY RUN completado: %d páginas con textLayer recolectado", len(paginas_texto))
        _log_resumen(st, dry_run=True)
        return []

    # ── FASE A: llamadas API en paralelo ──
    def _procesar_pagina(idx, page_id, texto):
        """Trocea la página si es grande y envía cada chunk a la API secuencialmente.
        Acumula avisos y tokens de todos los chunks. El dedup por ROL-año en
        _filtrar_avisos (Filtro 6) colapsa los duplicados del solape."""
        chunks = _trocear_pagina(texto)
        all_avisos = []
        total_usage = {"in": 0, "out": 0}

        for ci, chunk in enumerate(chunks):
            chunk_label = f"{page_id}.{ci}" if len(chunks) > 1 else page_id
            try:
                avisos_raw, usage = _enviar_texto_a_claude(chunk_label, chunk)
            except Exception as e:
                log.error("Chunk %s falló tras reintentos: %s", chunk_label, e)
                avisos_raw = []
                usage = {"in": 0, "out": 0}
            all_avisos.extend(avisos_raw)
            total_usage["in"] += usage["in"]
            total_usage["out"] += usage["out"]

        return idx, page_id, all_avisos, total_usage

    resultados = [None] * len(paginas_texto)
    log.info("[Paso 7/8] Procesando %d páginas con %d workers concurrentes",
             len(paginas_texto), MAX_WORKERS)
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futuros = [
            executor.submit(_procesar_pagina, idx, page_id, texto)
            for idx, (page_id, texto) in enumerate(paginas_texto)
        ]
        for fut in futuros:
            idx, page_id, avisos_raw, usage = fut.result()
            resultados[idx] = (page_id, avisos_raw, usage)
            st.tokens_input += usage["in"]
            st.tokens_output += usage["out"]

    # ── FASE B: post-proceso secuencial en orden original ──
    avisos_normalizados_total: list[dict[str, Any]] = []
    for i, (page_id, avisos_raw, _usage) in enumerate(resultados):
        log.info("Post-procesando página %d/%d: %s (%d avisos raw)",
                 i + 1, len(resultados), page_id, len(avisos_raw))
        st.avisos_vision += len(avisos_raw)
        for raw in avisos_raw:
            aviso_normalizado = _normalizar_aviso(raw)
            if aviso_normalizado is not None:
                cbr = _evaluar_cbr_por_anio(aviso_normalizado.get("año_inscripcion_dominio"))
                if cbr["decision"] == "EXCLUIR":
                    log.info("Aviso ROL %s descartado por CBR: dominio %s >= 2020",
                             aviso_normalizado.get("rol"), cbr["cbr_anio"])
                    st.excluidos_cbr.append({
                        "tribunal": aviso_normalizado.get("tribunal"),
                        "rol": aviso_normalizado.get("rol"),
                        "año": aviso_normalizado.get("año"),
                        "cbr_anio": cbr["cbr_anio"],
                    })
                    continue
                aviso_normalizado["cbr_anio"] = cbr["cbr_anio"]
                aviso_normalizado["cbr_flag_revision"] = cbr["cbr_flag_revision"]
                aviso_normalizado["cbr_motivo"] = cbr["cbr_motivo"]
                avisos_normalizados_total.append(aviso_normalizado)

    log.info("[Paso 8/8] Aplicando filtros a %d avisos normalizados", len(avisos_normalizados_total))
    todas_las_causas = _filtrar_avisos(avisos_normalizados_total, historico, vistos_en_ejecucion, st=st)
    st.avisos_post_filtro = len(todas_las_causas)
    st.causas_nuevas = len(todas_las_causas)

    log.info("=== Extracción completada: %d causas nuevas ===", len(todas_las_causas))
    _log_resumen(st)

    # Poblar métricas accesibles desde el exterior (canal lateral para M5)
    global _ultimas_metricas
    # Deduplicar excluidos CBR por ROL (primera aparición)
    _vistos = set()
    _exc_unicos = []
    for e in st.excluidos_cbr:
        if e["rol"] not in _vistos:
            _vistos.add(e["rol"])
            _exc_unicos.append(e)
    # Deduplicar descartados por ROL
    _des_vistos = set()
    _des_unicos = []
    for d in st.descartados:
        if d["rol"] not in _des_vistos:
            _des_vistos.add(d["rol"])
            _des_unicos.append(d)
    _ultimas_metricas = {
        "tokens_input": st.tokens_input,
        "tokens_output": st.tokens_output,
        "excluidos_cbr": _exc_unicos,
        "descartados": _des_unicos,
    }

    return todas_las_causas


# Canal lateral de métricas para M5 (no toca la firma de extraer_mercurio)
_ultimas_metricas: dict[str, Any] = {}


def obtener_metricas():
    """Retorna dict con tokens_input, tokens_output, excluidos_cbr de la última ejecución."""
    return _ultimas_metricas


# ===========================================================================
# API PÚBLICA
# ===========================================================================

def extraer_mercurio(
    fecha: date | str | None = None,
    *,
    dry_run: bool = False,
) -> list[dict[str, Any]]:
    """
    Extrae avisos de remates judiciales de propiedades (sección 1616) desde
    El Mercurio Digital para la fecha indicada (por defecto: hoy).

    Parámetros
    ----------
    fecha : date | str | None
        Fecha de la edición a procesar. Acepta:
        - None → hoy (date.today())
        - date object
        - str en formato "YYYY-MM-DD"
    dry_run : bool
        Si True, ejecuta solo navegación y lectura de textLayer (sin Claude API).

    Retorna
    -------
    list[dict]
        Lista de causas con las claves del contrato de datos:
        rol, año, corte, tribunal, demandante, demandado,
        direccion, comuna, region_rm (siempre True)
    """
    if fecha is None:
        fecha_obj = date.today()
    elif isinstance(fecha, str):
        fecha_obj = datetime.strptime(fecha, "%Y-%m-%d").date()
    else:
        fecha_obj = fecha

    return asyncio.run(_extraer_mercurio_async(fecha_obj, dry_run=dry_run))


# ===========================================================================
# CLI: permite ejecutar directamente  python modulo1_mercurio.py [--fecha YYYY-MM-DD]
# ===========================================================================

if __name__ == "__main__":
    import argparse, pprint

    parser = argparse.ArgumentParser(
        description="Extractor El Mercurio Digital — sección 1616 Remates de propiedades"
    )
    parser.add_argument(
        "--fecha",
        type=str,
        default=None,
        help="Fecha de la edición a procesar (YYYY-MM-DD). Por defecto: hoy.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Solo navegación y lectura de textLayer (sin llamar a Claude API).",
    )
    args = parser.parse_args()

    causas = extraer_mercurio(fecha=args.fecha, dry_run=args.dry_run)
    print(f"\n{'='*60}")
    print(f"CAUSAS EXTRAÍDAS: {len(causas)}")
    print("="*60)
    pprint.pprint(causas, width=120)
