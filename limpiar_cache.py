"""
limpiar_cache.py — Limpia el estado de corridas anteriores antes de una nueva prueba.

Uso:
    python limpiar_cache.py

Acciones:
    1. Borra todos los archivos de Descargas\\
    2. Elimina las filas de datos de la hoja CAUSAS en causas_ojv.xlsx
       (deja solo el encabezado de columnas intacto)

Por qué es necesario antes de cada prueba:
    - Descargas/ puede contener PDFs de corridas anteriores (incluyendo archivos
      con contenido incorrecto por el bug de "primera fila" corregido en v10.1).
    - La hoja CAUSAS acumula el historial de ROLes procesados. Si no se limpia,
      M1 los marca como duplicados y devuelve 0 causas nuevas al pipeline.
"""

import os
import sys
import glob
import shutil

import openpyxl

from config import DESCARGAS_DIR, CAUSAS_XLSX, SHEET_CAUSAS


# ─────────────────────────────────────────────────────────────────
# Helpers de presentación
# ─────────────────────────────────────────────────────────────────

def _contar_descargas():
    archivos = glob.glob(os.path.join(DESCARGAS_DIR, "*"))
    return len(archivos)

def _contar_filas_causas():
    try:
        wb = openpyxl.load_workbook(CAUSAS_XLSX)
        ws = wb[SHEET_CAUSAS]
        # max_row incluye el header; filas de datos = max_row - 1
        filas = ws.max_row - 1
        wb.close()
        return max(filas, 0)
    except Exception:
        return 0


def _limpiar_descargas():
    """Borra todos los archivos dentro de DESCARGAS_DIR (no borra el directorio)."""
    archivos = glob.glob(os.path.join(DESCARGAS_DIR, "*"))
    if not archivos:
        print("  Descargas/ ya está vacía.")
        return 0
    for ruta in archivos:
        try:
            if os.path.isfile(ruta):
                os.remove(ruta)
            elif os.path.isdir(ruta):
                shutil.rmtree(ruta)
        except Exception as e:
            print(f"  ERROR borrando {ruta}: {e}")
    print(f"  {len(archivos)} archivo(s) eliminado(s) de Descargas/")
    return len(archivos)


def _limpiar_causas_xlsx():
    """Elimina todas las filas de datos de la hoja CAUSAS, deja solo el header."""
    try:
        wb = openpyxl.load_workbook(CAUSAS_XLSX)
    except FileNotFoundError:
        print(f"  AVISO: {CAUSAS_XLSX} no encontrado — omitiendo.")
        return 0

    if SHEET_CAUSAS not in wb.sheetnames:
        print(f"  AVISO: hoja '{SHEET_CAUSAS}' no existe en el Excel — omitiendo.")
        wb.close()
        return 0

    ws = wb[SHEET_CAUSAS]
    filas_antes = ws.max_row - 1   # sin contar header

    # Borrar desde fila 2 hasta el final (fila 1 = header)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    wb.save(CAUSAS_XLSX)
    wb.close()

    print(f"  {max(filas_antes, 0)} fila(s) de datos eliminada(s) de hoja '{SHEET_CAUSAS}'")
    return max(filas_antes, 0)


# ─────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────

def main():
    print()
    print("=" * 55)
    print("  LIMPIAR CACHE — Sistema de Remates Judiciales")
    print("=" * 55)

    n_desc   = _contar_descargas()
    n_causas = _contar_filas_causas()

    print()
    print("  Se realizarán las siguientes acciones:")
    print(f"    1. Borrar {n_desc} archivo(s) en Descargas/")
    print(f"    2. Eliminar {n_causas} fila(s) de datos en hoja CAUSAS")
    print()

    if n_desc == 0 and n_causas == 0:
        print("  Nada que limpiar. Todo ya está vacío.")
        print()
        return

    respuesta = input("  ¿Confirmar? (s/n): ").strip().lower()
    if respuesta not in ("s", "si", "sí", "y", "yes"):
        print()
        print("  Cancelado.")
        print()
        sys.exit(0)

    print()
    _limpiar_descargas()
    _limpiar_causas_xlsx()

    print()
    print("  Listo. Puedes correr: python main.py --hasta 2")
    print("=" * 55)
    print()


if __name__ == "__main__":
    main()
